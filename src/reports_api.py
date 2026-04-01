import flask
from flask import Blueprint, abort, send_file
from db_helper import query_db
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import io
reports_bp = Blueprint('reports', __name__)
#Return
def success_response(data):
    return flask.jsonify({
        "success": True,
        "count": len(data) if isinstance(data, list) else 1,
        "data": data
    }), 200

#Ton Kho
@reports_bp.route('', methods=['GET'])
def getTonKho():
    """
    API lấy danh sách tồn kho
    ---
    tags:
      - Reports
    responses:
      200:
        description: Danh sách tồn kho
      404:
        description: Không có sản phẩm nào trong kho
    """
    query = """
                SELECT i.id, w.name AS warehouse_name, p.name AS product_name, i.quantity
                FROM Inventory i
                         JOIN Products p ON i.product_id = p.id
                         JOIN Warehouses w ON i.warehouse_id = w.id
                WHERE i.quantity > 0 
            """
    results = query_db(query)
    if results:
        return success_response(results)
    else:
        abort(404, description="Không có sản phẩm nào trong kho")


#Low-stock
@reports_bp.route('/low-stock', methods=['GET'])
def getLowStock():
    """
    API sản phẩm sắp hết hàng
    ---
    tags:
      - Reports
    responses:
      200:
        description: Danh sách sản phẩm sắp hết
      400:
        description: Không có sản phẩm nào sắp hết hàng
    """
    query = """
                SELECT p.name AS product_name, p.min_stock,
                       ISNULL(SUM(i.quantity), 0) AS total_quantity,
                       (p.min_stock - ISNULL(SUM(i.quantity), 0)) AS need_to_import
                FROM Products p
                         LEFT JOIN Inventory i ON p.id = i.product_id
                GROUP BY p.name, p.min_stock
                HAVING ISNULL(SUM(i.quantity), 0) <= p.min_stock 
            """
    results = query_db(query)
    if results:
        return success_response(results)
    else:
        abort(404, description="Không có sản phẩm nào sắp hết hàng")


#History
@reports_bp.route('/history/<string:sku>', methods=['GET'])
def getInventoryHistory(sku):
    """
    API lịch sử nhập/xuất
    ---
    tags:
      - Reports
    parameters:
      - name: sku
        in: path
        required: true
        type: string
    responses:
      200:
        description: Lịch sử nhập/xuất
      404:
        description: Không tìm thấy lịch sử nhập/xuất nào cho sản phẩm này
    """
    query = """
                SELECT p.sku, p.name AS product_name,
                       w.name AS warehouse_name,
                       l.change_amount, l.action_type, l.created_at
                FROM Inventory_Logs l
                         JOIN Products p ON l.product_id = p.id
                         LEFT JOIN Warehouses w ON l.warehouse_id = w.id
                WHERE p.sku = ?
                ORDER BY l.created_at DESC 
            """
    results = query_db(query, (sku,))
    if results:
        return success_response(results)
    else:
        abort(404, description="Không tìm thấy lịch sử nhập/xuất nào cho sản phẩm này")


#Transfer
@reports_bp.route('/transfer-history', methods=['GET'])
def getTransferHistory():
    """
    API lịch sử điều chuyển
    ---
    tags:
      - Reports
    parameters:
      - name: sku
        in: query
        required: false
        type: string
    responses:
      200:
        description: Lịch sử điều chuyển
      404:
        description: Không tìm thấy lịch sử điều chuyển nào cho sản phẩm này
    """
    sku = flask.request.args.get('sku')
    query = """
                SELECT p.sku, p.name AS product_name,
                       w_from.name AS from_warehouse,
                       w_to.name AS to_warehouse,
                       td.quantity,
                       ISNULL(u.username, 'System') AS staff,
                       t.status, t.created_at
                FROM Transfer_Details td
                         JOIN Transfer_Orders t ON td.transfer_id = t.id
                         JOIN Products p ON td.product_id = p.id
                         JOIN Warehouses w_from ON t.from_warehouse_id = w_from.id
                         JOIN Warehouses w_to ON t.to_warehouse_id = w_to.id
                         LEFT JOIN Users u ON t.staff_id = u.id
            """
    params = []
    if sku:
        query += " WHERE p.sku = ?"
        params.append(sku)
    query += " ORDER BY t.created_at DESC"
    results = query_db(query, tuple(params))
    if results:    
        return success_response(results)
    else:
        abort(404, description="Không tìm thấy lịch sử điều chuyển nào cho sản phẩm này")


#Receipts
@reports_bp.route('/receipt', methods=['GET'])
def getReceiptHistory():
    """
    API lịch sử nhập/xuất (Receipts)
    ---
    tags:
      - Reports
    responses:
      200:
        description: Lịch sử phiếu nhập/xuất
      404:
        description: Không tìm thấy lịch sử phiếu nhập/xuất nào
    """
    query = """
                SELECT r.id AS receipt_id, r.type,
                       w.name AS warehouse_name,
                       p.sku, p.name AS product_name,
                       rd.quantity, rd.price,
                       (rd.quantity * rd.price) AS total_value,
                       ISNULL(u.username, 'System') AS staff,
                       r.partner_name, r.created_at
                FROM Receipts r
                         JOIN Receipt_Details rd ON r.id = rd.receipt_id
                         JOIN Products p ON rd.product_id = p.id
                         JOIN Warehouses w ON r.warehouse_id = w.id
                         LEFT JOIN Users u ON r.staff_id = u.id
                ORDER BY r.created_at DESC 
            """
    results = query_db(query)
    if results:
        return success_response(results)
    else:
        abort(404, description="Không tìm thấy lịch sử phiếu nhập/xuất nào")


#Export
@reports_bp.route('/export/excel', methods=['GET'])
def export():
    """
    API export Excel
    ---
    tags:
      - Reports
    produces:
      - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    responses:
      200:
        description: File Excel
    """
    output = io.BytesIO()
    df1 = pd.DataFrame(query_db("""
                                    SELECT w.name AS warehouse, p.sku, p.name AS product, i.quantity
                                    FROM Inventory i
                                             JOIN Products p ON i.product_id = p.id
                                             JOIN Warehouses w ON i.warehouse_id = w.id
                                """))

    df2 = pd.DataFrame(query_db("""
                                    SELECT p.name, p.min_stock,
                                           ISNULL(SUM(i.quantity),0) total
                                    FROM Products p
                                             LEFT JOIN Inventory i ON p.id=i.product_id
                                    GROUP BY p.name,p.min_stock
                                """))

    df3 = pd.DataFrame(query_db("""
                                    SELECT TOP 100 p.sku,p.name,l.change_amount,l.action_type,l.created_at
                                    FROM Inventory_Logs l
                                             JOIN Products p ON l.product_id=p.id
                                    ORDER BY l.created_at DESC
                                """))

    df4 = pd.DataFrame(query_db("""
                                    SELECT p.sku,p.name,w_from.name AS from_wh,w_to.name AS to_wh,
                                           td.quantity,t.status,t.created_at
                                    FROM Transfer_Details td
                                             JOIN Transfer_Orders t ON td.transfer_id=t.id
                                             JOIN Products p ON td.product_id=p.id
                                             JOIN Warehouses w_from ON t.from_warehouse_id=w_from.id
                                             JOIN Warehouses w_to ON t.to_warehouse_id=w_to.id
                                """))

    df5 = pd.DataFrame(query_db("""
                                    SELECT r.id AS receipt_id, r.type,
                       w.name AS warehouse_name,
                       p.sku, p.name AS product_name,
                       rd.quantity, rd.price,
                       (rd.quantity * rd.price) AS total_value,
                       ISNULL(u.username, 'System') AS staff,
                       r.partner_name, r.created_at
                FROM Receipts r
                         JOIN Receipt_Details rd ON r.id = rd.receipt_id
                         JOIN Products p ON rd.product_id = p.id
                         JOIN Warehouses w ON r.warehouse_id = w.id
                         LEFT JOIN Users u ON r.staff_id = u.id
                ORDER BY r.created_at DESC
                                """))
    if not df5.empty:
        total_value_sum = df5['total_value'].sum()
        total_row = {col: '' for col in df5.columns}
        total_row['product_name'] = 'TOTAL'
        total_row['total_value'] = total_value_sum
        df5 = pd.concat([df5, pd.DataFrame([total_row])], ignore_index=True)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_list = {
                'TonKho': df1,
                'LowStock': df2,
                'History': df3,
                'Transfer': df4,
                'Receipts': df5
            }
            for sheet_name, df in df_list.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                for col_idx, col in enumerate(df.columns, 1):
                    max_length = len(str(col))
                    for cell in df[col]:
                        if cell:
                            # format
                            if "created_at" in col.lower():
                                cell = str(cell)[:19]
                            max_length = max(max_length, len(str(cell)))
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = max_length + 4
                    for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.alignment = Alignment(
                                wrap_text=True,
                                vertical='center'
                            )
    output.seek(0)
    return send_file(
            output,
            download_name="inventory_report.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )